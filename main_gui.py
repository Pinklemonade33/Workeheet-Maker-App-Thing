import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os
import pickle
from tkinter import filedialog
from win32com.client import Dispatch


def get_treeview_values(treeview, col=None):
    r_list = []
    if col is None:
        for a in treeview.get_children():
            value = treeview.item(a)['values']
            r_list.append(value)

    else:
        for a in treeview.get_children():
            value = treeview.item(a)['values'][col]
            r_list.append(value)

    return r_list


def clear_tree(treeview, index):
    for row in range(index):
        treeview.delete(row)


def add_non_duplicates_menu(menu, value, command, index):
    for x in range(index + 1):
        if menu.entrycget(x, 'label') == value:
            break
    else:
        menu.add_command(label=value, command=command)


def browse_files():
    filename = filedialog.askopenfilename(initialdir='Downloads',
                                          defaultextension='.xlsx',
                                          filetypes=[('Excel file', '.xlsx'), ('All files', '.*')])
    return filename


def find_total_length(values):
    total_len = 0
    for x in values:
        total_len += len(x)

    return total_len


class App:
    def __init__(self):
        self.root = tk.Tk()
        self.root.call('source', r'Sun-Valley-ttk-theme-master\sun-valley.tcl')
        self.root.call('set_theme', 'dark')

        self.start_frame = ttk.Frame(self.root)
        self.start_frame.pack()
        self.dtob = ttk.Button(self.start_frame, text='DTOB', command=self.start_dtob)
        self.material_finder = ttk.Button(self.start_frame, text='Material Finder', command=self.start_material_finder)
        self.dtob.pack(pady=(30, 10), padx=5, fill='x')
        self.material_finder.pack(pady=(10, 30), padx=5, fill='x')

        self.page = None

        self.root.mainloop()

    def start_dtob(self):
        self.start_frame.pack_forget()
        self.page = DtobPage1(self)

    def start_material_finder(self):
        self.start_frame.pack_forget()
        self.page = MfinderPage1(self)


class MfinderPage1:
    def __init__(self, app):
        self.app = app
        self.import_index_1 = 0
        self.import_index_2 = 0
        self.site_material = {}
        self.quantity = None
        self.path = None

        try:
            self.categories = pickle.load(open('pickles/categories.p', 'rb'))
        except FileNotFoundError:
            self.categories = {}

        try:
            self.translations_direct = pickle.load(open('pickles/translations_direct.p', 'rb'))
        except FileNotFoundError:
            self.translations_direct = {}

        try:
            self.translations_indirect = pickle.load(open('pickles/translations_indirect.p', 'rb'))
        except FileNotFoundError:
            self.translations_indirect = {}

        self.selected_category = tk.StringVar()
        self.selected_category.set('Select Category')
        self.selected_sites = []

        self.main_frame = ttk.Frame(app.root)
        self.tree_frame = ttk.Frame(self.main_frame)
        self.top_frame = ttk.Frame(self.main_frame)
        self.left_frame = ttk.Frame(self.main_frame)
        self.right_frame = ttk.Frame(self.main_frame, width=150)
        self.right_frame.grid_propagate(False)

        self.main_frame.grid_columnconfigure(0, weight=1)
        self.main_frame.grid_rowconfigure(1, weight=1)
        self.tree_frame.grid_columnconfigure(0, weight=1)
        self.tree_frame.grid_columnconfigure(1, weight=1)
        self.tree_frame.grid_rowconfigure(0, weight=1)

        self.main_frame.pack(fill=tk.BOTH, expand=True)
        self.top_frame.grid(row=0, column=0, sticky=tk.W + tk.E)
        self.tree_frame.grid(row=1, column=0, sticky=tk.N + tk.S + tk.E + tk.W)

        self.import_button = ttk.Button(self.top_frame, text='Import', command=self.import_tree)
        self.category_select = ttk.OptionMenu(self.top_frame, self.selected_category,
                                              self.selected_category.get(),
                                              *self.categories.keys(),
                                              command=lambda args: self.set_category())
        self.print_button = ttk.Button(self.top_frame, text='Review and Print', command=self.print_popup)
        self.translate_name_btn = ttk.Button(self.top_frame, text='Translate', command=self.translate_popup)
        self.categories_edit = ttk.Button(self.top_frame, text='Edit Categories', command=self.popup_category_win)

        self.print_var = tk.IntVar()
        self.print_var.set(0)
        self.print_only_selected = ttk.Checkbutton(self.top_frame, text='Print Only Selected',
                                                   variable=self.print_var, onvalue=0, offvalue=1,
                                                   style='Switch.TCheckbutton')
        self.translation_var = tk.IntVar()
        self.translation_var.set(0)
        self.translate_check = ttk.Checkbutton(self.top_frame, text=' Show Translations',
                                               variable=self.translation_var, command=lambda: self.reload(),
                                               onvalue=0, offvalue=1, style='Switch.TCheckbutton')

        self.show_selected_var = tk.IntVar()
        self.show_selected_var.set(1)
        self.show_only_selected_check = ttk.Checkbutton(self.top_frame, text='Show Only Selected',
                                                        variable=self.show_selected_var,
                                                        command=lambda: self.only_selected_on_off(),
                                                        onvalue=0, offvalue=1, style='Switch.TCheckbutton')

        self.import_button.grid(row=0, column=0, padx=5, pady=5)
        self.category_select.grid(row=0, column=1, padx=5, pady=5)
        self.top_frame.columnconfigure(4, weight=1)
        self.print_button.grid(row=0, column=2, padx=5, pady=5)
        self.translate_name_btn.grid(row=0, column=3, padx=5, pady=5, sticky=tk.W)
        self.categories_edit.grid(row=0, column=5, padx=5, pady=5, sticky=tk.E)

        self.print_only_selected.grid(row=1, column=0, padx=5, pady=5, sticky=tk.W, columnspan=2)
        self.translate_check.grid(row=1, column=2, pady=5, padx=5, sticky=tk.W, columnspan=2)
        self.show_only_selected_check.grid(row=1, column=4, pady=5, padx=5, sticky=tk.W, columnspan=2)

        self.import_tree_1 = ttk.Treeview(self.tree_frame)
        self.import_tree_1.bind('<ButtonRelease-1>', lambda args: self.load_tree_2())
        self.import_tree_1.column('#0', width=0, stretch=False)
        self.import_tree_1['columns'] = [0, 1]
        self.import_tree_1.column(0, width=120)
        self.import_tree_1.column(1, width=120)
        self.import_tree_1.heading(0, text='Job#')
        self.import_tree_1.heading(1, text='Site ID')

        self.import_tree_2 = ttk.Treeview(self.tree_frame)
        self.import_tree_2.column('#0', width=0, stretch=False)
        self.import_tree_2['columns'] = [0, 1, 2]
        self.import_tree_2.column(0, width=100)
        self.import_tree_2.column(1, width=220)
        self.import_tree_2.column(2, width=70, stretch=False)
        self.import_tree_2.heading(0, text='Material')
        self.import_tree_2.heading(1, text='Description')
        self.import_tree_2.heading(2, text='Quantity', command=lambda: self.sort_by_quantity())
        self.import_tree_2.bind('<Double-ButtonRelease-1>', lambda args: self.context_popup_1(args))

        self.import_tree_2.tag_configure('1', background='dark olive green')
        self.import_tree_1.tag_configure('1', background='dark olive green')

        self.import_tree_1.grid(row=0, column=0, sticky=tk.N + tk.S + tk.E + tk.W)
        self.import_tree_2.grid(row=0, column=1, rowspan=2, sticky=tk.N + tk.S + tk.E + tk.W)

        self.context_menu = tk.Menu(self.app.root, tearoff=0)
        self.context_menu.add_command(label='Add to Category',
                                      command=lambda: self.add_category_item())
        self.context_menu.add_command(label='Remove from Category',
                                      command=lambda: self.remove_category_item())

    def only_selected_on_off(self):
        def show_only_selected():
            clear_tree(self.import_tree_1, self.import_index_1)
            clear_tree(self.import_tree_2, self.import_index_2)
            self.import_index_1 = 0
            self.import_index_2 = 0
            for site in self.selected_sites:
                value = self.site_material[site]
                self.import_tree_1.insert('', 'end', str(self.import_index_1), values=(site, value[0][0]))
                self.import_index_1 += 1

            self.highlight_sites()

        def show_all():
            for key, value in self.site_material.items():
                self.import_tree_1.insert('', 'end', str(self.import_index_1), values=(key, value[0][0]))
                self.import_index_1 += 1

        if self.show_selected_var.get() == 0:
            show_only_selected()
        else:
            show_all()

    def translate_popup(self):
        def enter_translation_direct():
            nonlocal tree_direct_index
            tree_direct.insert('', 'end', str(tree_direct_index),
                               values=[text_entry_direct.get(),
                               translation_entry_direct.get()])
            tree_direct_index += 1
            translation_changes_direct.update({text_entry_direct.get(): translation_entry_direct.get()})

        def enter_translation_indirect():
            nonlocal tree_indirect_index
            if text_entry_indirect_2.get() == '':
                val_in_2 = 0
            else:
                val_in_2 = text_entry_indirect_2.get()
            if text_entry_indirect_3.get() == 'Same':
                val_in_3 = text_entry_indirect_1.get()
            else:

                val_in_3 = text_entry_indirect_3.get()
            if text_entry_indirect_4.get() == '':
                val_in_4 = 0
            else:
                val_in_4 = text_entry_indirect_4.get()
            tree_indirect.insert('', 'end', str(tree_indirect_index),
                                 values=[text_entry_indirect_1.get(),
                                 val_in_2,
                                 val_in_3,
                                 val_in_4])
            tree_indirect_index += 1
            translation_changes_indirect.update({
                text_entry_indirect_1.get():
                    [val_in_2,
                     val_in_3,
                     val_in_4]})

        def remove_translation_direct():
            if len(tree_direct.selection()) != 0:
                index = tree_direct.selection()[0]
                dict_value = tree_direct.item(index)['values'][0]
                translation_changes_direct.pop(dict_value)
                tree_direct.delete(index)

        def remove_translation_indirect():
            if len(tree_indirect.selection()) != 0:
                index = tree_indirect.selection()[0]
                dict_value = tree_indirect.item(index)['values'][0]
                translation_changes_indirect.pop(dict_value)
                tree_indirect.delete(index)

        def save_translation():
            def update_material():
                for site in self.site_material.keys():
                    for row_index, row in enumerate(self.site_material[site]):
                        val = self.translate(row)
                        self.site_material[site][row_index] = val

                self.reload()

            self.translations_direct = translation_changes_direct
            self.translations_indirect = translation_changes_indirect
            pickle.dump(self.translations_direct, open('pickles/translations_direct.p', 'wb'))
            pickle.dump(self.translations_indirect, open('pickles/translations_indirect.p', 'wb'))
            if self.path is not None:
                update_material()
            win.destroy()

        win = tk.Toplevel(self.app.root)
        note = ttk.Notebook(win)
        note_1 = ttk.Frame(note)
        note_2 = ttk.Frame(note)
        note.add(note_1, text='Direct Translation')
        note.add(note_2, text='Indirect Translation')
        tree_direct_index = 0
        tree_indirect_index = 0
        translation_changes_direct = self.translations_direct
        translation_changes_indirect = self.translations_indirect

        save_btn = ttk.Button(win, text='Save', command=save_translation)

        tree_direct = ttk.Treeview(note_1)
        tree_direct.bind('<Double-Button-1>', lambda args: remove_translation_direct())
        tree_direct.column('#0', width=0, stretch=False)
        tree_direct['columns'] = [0, 1]
        tree_direct.heading(0, text='Original text')
        tree_direct.heading(1, text='Direct Translation')
        for key, value in translation_changes_direct.items():
            tree_direct.insert('', 'end', str(tree_direct_index), values=[key, value])
            tree_direct_index += 1

        tree_indirect = ttk.Treeview(note_2)
        tree_indirect.bind('<Double-Button-1>', lambda args: remove_translation_indirect())
        tree_indirect.column('#0', width=0, stretch=False)
        tree_indirect['columns'] = [0, 1, 2, 3]
        tree_indirect.column(1, width=5)
        tree_indirect.column(2, anchor=tk.CENTER)
        tree_indirect.column(3, width=5)
        tree_indirect.heading(0, text='Original text')
        tree_indirect.heading(1, text='-')
        tree_indirect.heading(2, text='Translation')
        tree_indirect.heading(3, text='+')
        for key, value in translation_changes_indirect.items():
            tree_indirect.insert('', 'end', str(tree_indirect_index), values=[key, value[0], value[1], value[2]])
            tree_indirect_index += 1

        text_entry_direct = ttk.Entry(note_1)
        to_label_direct = ttk.Label(note_1, text='=')
        translation_entry_direct = ttk.Entry(note_1)
        enter_btn_direct = ttk.Button(note_1, text='Enter', command=enter_translation_direct)

        text_entry_indirect_1 = ttk.Entry(note_2)
        text_entry_indirect_2 = ttk.Entry(note_2, width=1)
        text_entry_indirect_3 = ttk.Entry(note_2, width=10)
        text_entry_indirect_3.insert('end', 'Same')
        text_entry_indirect_4 = ttk.Entry(note_2, width=1)
        to_label_indirect_1 = ttk.Label(note_2, text='=')
        to_label_indirect_2 = ttk.Label(note_2, text=' + ')
        to_label_indirect_4 = ttk.Label(note_2, text=' Back ')
        to_label_indirect_5 = ttk.Label(note_2, text=' Forward ')
        enter_btn_indirect = ttk.Button(note_2, text='Enter', command=enter_translation_indirect)

        text_entry_direct.grid(row=0, column=0, pady=(5, 4), padx=3)
        to_label_direct.grid(row=0, column=1, pady=(5, 4), padx=3)
        translation_entry_direct.grid(row=0, column=2, pady=(5, 4), padx=3)
        enter_btn_direct.grid(row=0, column=3, pady=(5, 4), padx=3, sticky=tk.E + tk.W)

        text_entry_indirect_1.grid(row=0, column=0, pady=(5, 4), padx=(3, 1))
        text_entry_indirect_2.grid(row=0, column=4, pady=(5, 4), padx=1)
        text_entry_indirect_4.grid(row=0, column=6, pady=(5, 4), padx=1)
        text_entry_indirect_3.grid(row=0, column=2, pady=(5, 4), padx=1)
        to_label_indirect_1.grid(row=0, column=1, pady=(5, 4), padx=1)
        to_label_indirect_2.grid(row=0, column=3, pady=(5, 4), padx=1)
        to_label_indirect_4.grid(row=0, column=5, pady=(5, 4), padx=1)
        to_label_indirect_5.grid(row=0, column=7, pady=(5, 4), padx=1)
        enter_btn_indirect.grid(row=0, column=9, pady=(5, 4), padx=(1, 3))

        note.pack()
        note_1.columnconfigure(3, weight=1)
        tree_direct.grid(row=1, column=0, columnspan=4, sticky=tk.E + tk.W, padx=3)
        tree_indirect.grid(row=1, column=0, columnspan=10, sticky=tk.E + tk.W, padx=3)

        save_btn.pack(anchor=tk.SE, pady=3)

    def context_popup_1(self, event):
        if len(self.import_tree_2.selection()) != 0:
            self.context_menu.tk_popup(event.x_root, event.y_root, 0)

    def highlight_category_items(self):
        if self.selected_category.get() == 'Select Category':
            return
        else:
            category = self.categories[self.selected_category.get()]

        if len(category) > 0:
            for index in self.import_tree_2.get_children():
                if self.import_tree_2.item(index)['values'][0] in category:
                    self.import_tree_2.item(index, tags='1')
                else:
                    self.import_tree_2.item(index, tags='2')
        else:
            for index in self.import_tree_2.get_children():
                self.import_tree_2.item(index, tags='2')

    def highlight_sites(self):
        def highlight_tree_1(site_value):
            for row in self.import_tree_1.get_children():
                if self.import_tree_1.item(row)['values'][0] == site_value:
                    self.import_tree_1.item(row, tags='1')

        def highlight_tree_2(site_value):
            for row in self.import_tree_1.get_children():
                if self.import_tree_1.item(row)['values'][0] == site_value:
                    self.import_tree_1.item(row, tags='2')

        if self.selected_category.get() == 'Select Category':
            return
        else:
            category = self.categories[self.selected_category.get()]

        self.selected_sites = []
        for site in self.site_material.keys():
            for index in range(len(self.site_material[site])):
                broke = False
                for item in category:
                    if self.site_material[site][index][1] == item:
                        highlight_tree_1(site)
                        self.selected_sites.append(site)
                        broke = True
                        break
                if broke is True:
                    break
            else:
                highlight_tree_2(site)

    def auto_size_tree_cols(self):
        # On Hold, not used
        highest_value = [0, 0, 0, 0, 0]
        sites = self.site_material.keys()
        for site in sites:
            if len(site) > highest_value[0]:
                highest_value[0] = len(site)

            if len(self.site_material[site][0][0]) > highest_value[1]:
                highest_value[1] = len(self.site_material[site][0][0])

            for row in self.site_material[site]:
                for index, col in enumerate(row):
                    if index > 0:
                        if len(str(col)) > highest_value[index + 1]:
                            highest_value[index + 1] = len(str(col))

        self.import_tree_1.column(0, width=highest_value[0])
        self.import_tree_1.column(1, width=highest_value[1])
        self.import_tree_2.column(0, width=highest_value[2])
        self.import_tree_2.column(1, width=highest_value[3])
        self.import_tree_2.column(2, width=highest_value[4])

    def import_tree(self, reload=False):
        def get_job_numbers():
            past = False
            for row in range(1, ws.max_row + 1):
                if past is True:
                    if ws.cell(row, 2).value is not None:
                        job_numbers.add(ws.cell(row, 2).value)

                if ws.cell(row, 2).value == 'Job #':
                    past = True

        def get_site_material():
            for job in job_numbers:
                job_data = []
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row, 2).value == job:
                        # 5=name, 8=item, 9=description, 13=quantity

                        data = [ws.cell(row, 5).value, ws.cell(row, 8).value,
                                ws.cell(row, 9).value, ws.cell(row, 14).value]

                        if self.translation_var.get() == 0:
                            data = self.translate(data)

                        job_data.append(data)
                self.site_material.update({job: job_data})

        def import_tree_1():
            for key, value in self.site_material.items():
                self.import_tree_1.insert('', 'end', str(self.import_index_1), values=(key, value[0][0]))
                self.import_index_1 += 1

        job_numbers = set()
        if reload is False:
            self.path = browse_files()
            if self.path[-4:] != 'xlsx':
                return

        wb = load_workbook(filename=self.path)
        ws = wb.active

        get_job_numbers()
        get_site_material()
        import_tree_1()
        self.highlight_sites()

    def load_tree_2(self):
        if len(self.import_tree_1.selection()):
            row = self.import_tree_1.selection()[0]
            sel_job = self.import_tree_1.item(row)['values'][0]
            clear_tree(self.import_tree_2, self.import_index_2)
            self.import_index_2 = 0

            for line in self.site_material[sel_job]:
                if self.show_selected_var.get() == 1:
                    self.import_tree_2.insert('', 'end', str(self.import_index_2), values=(line[1], line[2], line[3]))
                    self.import_index_2 += 1
                else:
                    if line[1] in self.categories[self.selected_category.get()]:
                        self.import_tree_2.insert('', 'end', str(self.import_index_2), values=(line[1], line[2], line[3]))
                        self.import_index_2 += 1

            self.highlight_category_items()

    def update_categories_select(self):
        for category in self.categories:
            self.category_select['menu'].add_command(label=category, command=lambda: self.set_category)

    def add_category_item(self):
        index = self.import_tree_2.selection()[0]
        value = self.import_tree_2.item(index)['values'][0]
        self.categories[self.selected_category.get()].add(value)
        self.set_category()

    def remove_category_item(self):
        index = self.import_tree_2.selection()[0]
        value = self.import_tree_2.item(index)['values'][0]
        self.categories[self.selected_category.get()].remove(value)
        self.set_category()

    def set_category(self):
        self.highlight_category_items()
        self.highlight_sites()
        pickle.dump(self.categories, open('pickles/categories.p', 'wb'))

    def popup_category_win(self):
        def set_category():
            value = var.get()
            nonlocal tree_index
            clear_tree(tree, tree_index)
            tree_index = 0

            for item in category_changes[value]:
                tree.insert('', 'end', str(tree_index), values=item)
                tree_index += 1

        def add_new_category():
            value = add_ent.get()
            category_changes.update({value: set()})
            add_non_duplicates_menu(option_menu['menu'], value,
                                    lambda: set_category(), len(category_changes.keys()))
            add_ent.delete(0, 'end')

        def delete_category():
            if var.get() != 'Select Category':
                nonlocal tree_index
                category_changes.pop(var.get())
                option_menu['menu'].delete(var.get())
                clear_tree(tree, tree_index)
                tree_index = 0
                var.set('Select Category')

        def save_changes():
            self.categories = category_changes
            self.category_select.destroy()
            self.category_select = ttk.OptionMenu(self.top_frame, self.selected_category,
                                                  self.selected_category.get(),
                                                  *self.categories.keys(),
                                                  command=lambda args: self.set_category())
            self.category_select.grid(row=0, column=1, padx=5, pady=5)
            pickle.dump(self.categories, open('pickles/categories.p', 'wb'))
            win.destroy()

        def context_popup_1(event):
            if var.get() != 'Select Category':
                if len(tree.selection()) != 0:
                    context_menu_1.tk_popup(event.x_root, event.y_root, 0)
                else:
                    context_menu_3.tk_popup(event.x_root, event.y_root, 0)

        def context_popup_2(event):
            if var.get() != 'Select Category':
                if len(tree.selection()) != 0:
                    context_menu_2.tk_popup(event.x_root, event.y_root, 0)

        def paste_clipboard():
            nonlocal tree_index
            value = self.app.root.clipboard_get()
            values = value.split()

            for val in values:
                tree.insert('', 'end', str(tree_index), values=val)
                tree_index += 1

        def remove_item():
            nonlocal tree_index
            index = tree.selection()[0]
            item = tree.item(index)['values'][0]
            category_changes[var.get()].remove(item)
            tree.delete(index)
            tree_index -= 1

        def clear_all():
            nonlocal tree_index
            category_changes[var.get()].clear()
            clear_tree(tree, tree_index)
            tree_index = 0

        def add_item():
            def add():
                nonlocal tree_index
                tree.insert('', 'end', str(tree_index), values=[text_ent.get(),])
                tree_index += 1
                category_changes[var.get()].add(text_ent.get())
                text_win.destroy()

            text_win = tk.Toplevel()
            text_win.bind('<Return>', lambda args: add())
            text_ent = ttk.Entry(text_win)
            text_btn = ttk.Button(text_win, text='Add', command=add)

            text_ent.grid(row=0, column=0)
            text_btn.grid(row=0, column=1)

        def change_name():
            def change():
                index = tree.selection()[0]
                item = tree.item(index)['values'][0]
                category_changes[var.get()].remove(item)
                category_changes[var.get()].add(text_ent.get())
                tree.delete(index)
                tree.insert('', int(index), str(index), values=[text_ent.get(), ])

            text_win = tk.Toplevel()
            text_win.bind('<Return>', lambda args: change())
            text_ent = ttk.Entry(text_win)
            text_btn = ttk.Button(text_win, text='Change', command=change)

            text_ent.grid(row=0, column=0)
            text_btn.grid(row=0, column=1)

        category_changes = self.categories
        tree_index = 0

        win = tk.Toplevel(self.app.root)

        top_frame = ttk.Frame(win)
        add_ent = ttk.Entry(top_frame)
        add_btn = ttk.Button(top_frame, text='Add', command=lambda: add_new_category())
        del_btn = ttk.Button(top_frame, text='Delete', command=delete_category)
        save_btn = ttk.Button(top_frame, text='Save', command=save_changes)

        var = tk.StringVar()
        var.set('Select Category')
        option_menu = ttk.OptionMenu(top_frame, var, var.get(), *self.categories.keys(),
                                     command=lambda args: set_category())

        tree = ttk.Treeview(top_frame, show='tree')
        tree.column('#0', width=0, stretch=False)
        tree['columns'] = [0]
        tree.bind('<Button-3>', context_popup_1)
        tree.bind('<Double-Button-1>', context_popup_2)

        context_menu_1 = tk.Menu(win, tearoff=0)
        context_menu_1.add_command(label='Add Item', command=add_item)
        context_menu_1.add_command(label='Paste', command=paste_clipboard)
        context_menu_1.add_command(label='Remove Item', command=remove_item)
        context_menu_1.add_command(label='Clear All', command=clear_all)
        context_menu_1.add_command(label='Change Name', command=change_name)

        context_menu_2 = tk.Menu(win, tearoff=0)
        context_menu_2.add_command(label='Remove Item', command=remove_item)
        context_menu_2.add_command(label='Change Name', command=change_name)

        context_menu_3 = tk.Menu(win, tearoff=0)
        context_menu_3.add_command(label='Add Item', command=add_item)
        context_menu_3.add_command(label='Paste', command=paste_clipboard)
        context_menu_3.add_command(label='Clear All', command=clear_all)

        top_frame.pack(fill='x', side=tk.TOP)
        add_ent.grid(row=0, column=0, sticky=tk.E + tk.W, pady=5, padx=5)
        add_btn.grid(row=0, column=2, sticky=tk.E + tk.W, pady=5, padx=5)
        option_menu.grid(row=1, column=0, columnspan=2, sticky=tk.E + tk.W, pady=5, padx=5)
        del_btn.grid(row=1, column=2, sticky=tk.E + tk.W, pady=5, padx=5)
        tree.grid(row=2, column=0, columnspan=3, sticky=tk.E + tk.W, pady=5, padx=5)
        save_btn.grid(row=3, column=2, sticky=tk.E + tk.W, pady=5, padx=5)

    def sort_by_quantity(self):
        new_list = []
        list_key = {}
        for index in self.import_tree_2.get_children():
            if len(list_key) == 0:
                list_key = {index: self.import_tree_2.item(index)['values'][2]}
            else:
                list_key.update({index: self.import_tree_2.item(index)['values'][2]})

            new_list.append(self.import_tree_2.item(index)['values'][2])
            self.import_tree_2.detach(index)

        if self.quantity == 'max' or self.quantity is None:
            new_list.sort()
            self.quantity = 'min'
        elif self.quantity == 'min':
            self.quantity = 'max'
            new_list.sort(reverse=True)

        for x in range(len(new_list)):
            for key, value in list_key.items():
                if value == new_list[x]:
                    list_key[key] = x
                    new_list[x] = None
                    break

        for index in range(len(new_list)):
            for key, value in list_key.items():
                if value == index:
                    self.import_tree_2.reattach(key, '', value)

    def print_popup(self):
        def load_sheet():
            def get_selected_material_by_material():

                def create_site_groups():
                    new_group = list()
                    group_count = 0
                    if op_var_2.get() == 2:
                        for site in self.selected_sites:

                            if group_count == div:
                                group_count = 0
                                groups.append(new_group)
                                new_group = list()

                            new_group.append(site)
                            group_count += 1

                        groups.append(new_group)
                    elif op_var_2.get() == 1:
                        for index in order_tree.get_children():
                            new_group = []
                            for site in order_tree.get_children(index):
                                if check_var.get() == 1:
                                    new_group.append(order_tree.item(site)['values'][0])
                                elif check_var.get() == 0:
                                    new_group.append(order_tree.item(site)['values'][1])

                            groups.append(new_group)

                def create_sections():
                    def create_new_section():
                        nonlocal section_count
                        nonlocal section_material_id
                        section_count = 0
                        new_section = list()
                        for x in section_material_id:
                            new_section.append(x)
                        new_section.sort()
                        new_section.insert(0, 'Job #')
                        new_section.insert(1, 'Site ID')
                        new_section_new_data = [new_section, []]
                        sections.append(new_section_new_data)
                        section_material_id = set()

                    current_ids = set()
                    section_material_id = set()
                    for group in groups:
                        sections = list()
                        section_count = 0
                        for site_name in group:
                            for row in self.site_material[site_name]:
                                if row[1] in self.categories[self.selected_category.get()]:
                                    if row[1] not in current_ids:
                                        section_material_id.add(row[1])
                                        current_ids.add(row[1])
                                        section_count += 1
                                        if section_count == column_max:
                                            create_new_section()

                        if section_count > 0:
                            create_new_section()
                        section_groups.append(sections)
                        current_ids = set()

                def create_section_data():
                    for x in range(len(section_groups)):
                        for section in section_groups[x]:

                            for site_name in groups[x]:
                                section_data = [site_name,
                                                self.site_material[site_name][0][0],
                                                " ", " ", " ", " ", " ", " "]
                                add = False
                                for row in self.site_material[site_name]:
                                    if row[1] in section[0]:
                                        add = True
                                        index = section[0].index(row[1])
                                        section_data[index] = row[3]

                                if add is True:
                                    section[1].append(section_data)

                def write_material():
                    def format_header():
                        nonlocal cell_row
                        nonlocal cell_col
                        nonlocal empty_rows
                        top_colors = PatternFill(patternType='solid', start_color='0070C0', end_color='0070C0')
                        top_font = Font(color='FFFFFF')

                        for col in range(1, cell_col):
                            ws.cell(cell_row, col).fill = top_colors
                            ws.cell(cell_row, col).font = top_font
                            ws.cell(cell_row, col).border = all_borders
                            ws.cell(cell_row, col).alignment = Alignment(wrap_text=True)

                        ws.row_dimensions[cell_row].height = 29
                        empty_rows -= 1

                    def format_columns():
                        ws.column_dimensions['A'].width = 15
                        ws.column_dimensions['B'].width = 15
                        ws.column_dimensions['C'].width = 15
                        ws.column_dimensions['D'].width = 15
                        ws.column_dimensions['E'].width = 15
                        ws.column_dimensions['F'].width = 15
                        ws.column_dimensions['G'].width = 15
                        ws.column_dimensions['H'].width = 15

                    def check_page_fit_1():
                        nonlocal empty_rows
                        nonlocal cell_row

                        if cell_row > 1:
                            cell_row += empty_rows
                            empty_rows = 32

                    def check_page_fit_2():
                        nonlocal empty_rows
                        nonlocal section
                        nonlocal cell_row
                        size = len(section[1]) + 2
                        if empty_rows - size < 0:
                            cell_row += empty_rows
                            empty_rows = 32

                    empty_rows = 32
                    all_borders = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                    cell_row = 0
                    format_columns()
                    nonlocal op_var
                    for section_group in section_groups:
                        if op_var.get() == 1:
                            check_page_fit_1()
                        for section in section_group:
                            if op_var.get() == 2:
                                check_page_fit_2()

                            cell_row += 1
                            empty_rows -= 1
                            cell_col = 1

                            for item in section[0]:
                                ws.cell(cell_row, cell_col).value = item
                                if item != " ":
                                    ws.cell(cell_row, cell_col).border = all_borders
                                cell_col += 1

                            format_header()
                            top_row = cell_row
                            for row in section[1]:
                                cell_row += 1
                                empty_rows -= 1
                                cell_col = 1
                                for item in row:
                                    ws.cell(cell_row, cell_col).value = item
                                    if ws.cell(top_row, cell_col).value is not None:
                                        ws.cell(cell_row, cell_col).border = all_borders
                                    cell_col += 1

                            empty_rows -= 2
                            cell_row += 2
                        empty_rows -= 1
                        cell_row += 1

                # 118.14 = max width for printing landscape
                # 460 = max height for printing landscape

                last_row = 32
                groups = list()
                section_groups = list()
                all_material_id = set()
                column_max = 6
                div = int(rows_per_group_entry.get())
                create_site_groups()
                create_sections()
                create_section_data()
                write_material()

            def get_selected_material_by_site():
                def get_material():
                    nonlocal export_data
                    data_group = []

                    for site in self.selected_sites:
                        row_group = [site, self.site_material[site][0][0], 'Section 1']
                        data_group.append(row_group)
                        row_group = ['Material ID', 'Material Description', 'Quantity']
                        data_group.append(row_group)
                        for row in self.site_material[site]:
                            if row[1] in self.categories[self.selected_category.get()]:
                                row_group = [row[1], row[2], row[3]]
                                data_group.append(row_group)

                        export_data.append(data_group)
                        data_group = []

                def write_material():
                    def format_header(row_index):
                        nonlocal cell_row
                        nonlocal empty_rows
                        top_colors_1 = PatternFill(patternType='solid', start_color='0070C0', end_color='0070C0')
                        top_colors_2 = PatternFill(patternType='solid', start_color='002060', end_color='002060')
                        top_font = Font(color='FFFFFF')

                        if row_index == 1:
                            top_colors = top_colors_1
                        elif row_index == 0:
                            top_colors = top_colors_2
                        else:
                            raise ValueError

                        for col in range(1, 4):
                            ws.cell(cell_row, col).fill = top_colors
                            ws.cell(cell_row, col).font = top_font
                            ws.cell(cell_row, col).border = all_borders
                            ws.cell(cell_row, col).alignment = Alignment(wrap_text=True)

                    def check_page_fit_1():
                        nonlocal empty_rows
                        nonlocal cell_row

                        if cell_row > 1:
                            cell_row += empty_rows
                            empty_rows = 32

                    def check_page_fit_2():
                        nonlocal empty_rows
                        nonlocal section
                        nonlocal cell_row
                        nonlocal section_count
                        if index == 0 and empty_rows - 4 < 0:
                            cell_row += empty_rows + 1
                            empty_rows = 31

                        if empty_rows == -1:
                            row_1 = section[0]
                            row_2 = section[1]
                            section_count += 1
                            ws.cell(cell_row, 1).value = row_1[0]
                            ws.cell(cell_row, 2).value = row_1[1]
                            ws.cell(cell_row, 3).value = 'Section ' + str(section_count)
                            format_header(0)
                            cell_row += 1
                            format_header(1)
                            ws.cell(cell_row, 1).value = row_2[0]
                            ws.cell(cell_row, 2).value = row_2[1]
                            ws.cell(cell_row, 3).value = row_2[2]
                            cell_row += 1

                            empty_rows = 30

                    ws.column_dimensions['A'].width = 40
                    ws.column_dimensions['B'].width = 70
                    ws.column_dimensions['C'].width = 10
                    all_borders = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))

                    empty_rows = 32
                    cell_row = 0
                    nonlocal op_var
                    for section in export_data:
                        section_count = 1
                        if op_var.get() == 1:
                            check_page_fit_1()
                        for index, row in enumerate(section):

                            cell_row += 1
                            empty_rows -= 1
                            if op_var.get() == 2:
                                check_page_fit_2()

                            if index == 0:
                                format_header(0)
                                ws.cell(cell_row, 1).value = row[0]
                                ws.cell(cell_row, 2).value = row[1]
                                ws.cell(cell_row, 3).value = row[2]

                            elif index == 1:
                                format_header(1)
                                ws.cell(cell_row, 1).value = row[0]
                                ws.cell(cell_row, 2).value = row[1]
                                ws.cell(cell_row, 3).value = row[2]

                            else:
                                ws.cell(cell_row, 1).value = row[0]
                                ws.cell(cell_row, 2).value = row[1]
                                ws.cell(cell_row, 3).value = row[2]
                                ws.cell(cell_row, 1).border = all_borders
                                ws.cell(cell_row, 2).border = all_borders
                                ws.cell(cell_row, 3).border = all_borders

                        cell_row += 1
                        empty_rows -= 1

                export_data = []
                get_material()
                write_material()

            if var.get() == 'By Site':
                get_selected_material_by_site()
            elif var.get() == 'By Material':
                get_selected_material_by_material()

            ws.set_printer_settings(paper_size=90, orientation='landscape')
            wb.save(filename=r'material.xlsx')
            os.startfile(r'material.xlsx')

        def reload_sheet():
            xl = Dispatch('Excel.Application')
            ws.delete_cols(1, ws.max_column)
            ws.delete_rows(1, ws.max_row)
            wb.close()
            dwb = xl.Workbooks.Open(r'C:\Users\braid\PycharmProjects\MacroProject\gui\material.xlsx')
            dwb.Close(True)

        def create_new_tree_section():
            nonlocal current_section
            nonlocal order_row_index

            if current_section == 0:
                current_section = 1

            if len(removed_sections) != 0:
                current_section = int(removed_sections[0])
                removed_sections.pop(0)

            new_section = 'Section ' + str(current_section)
            order_tree.insert('', current_section - 1,  str(current_section), values=[new_section])
            order_row_index.append(0)
            current_section = len(order_tree.get_children()) + 1

        def add_to_section():
            nonlocal site_tree_index
            if len(site_tree.selection()) == 0:
                return

            if current_section == 0:
                create_new_tree_section()
                order_tree.selection_set('1')
            elif len(order_tree.selection()) == 0:
                order_tree.selection_set('1')

            if ' ' in order_tree.selection()[0]:
                space_index = order_tree.selection()[0].index(' ')
                parent_index = str(order_tree.selection()[0][:space_index])
                index = order_row_index[int(parent_index) - 1]
            else:
                parent_index = str(order_tree.selection()[0])
                index = str(order_row_index[int(order_tree.selection()[0]) - 1])

            if check_var.get() == 1:
                val = site_tree.item(site_tree.selection()[0])['values']
            else:
                val = site_tree.item(site_tree.selection()[0])['values']
                val.reverse()

            order_tree.insert(parent_index,
                              'end',
                              parent_index + ' ' + str(index),
                              values=val)
            order_row_index[int(parent_index) - 1] += 1
            site_tree_removed.append(site_tree.selection()[0])
            site_tree.delete(site_tree.selection()[0])
            site_tree_index -= 1

        def remove_section():
            nonlocal site_tree_index
            nonlocal current_section
            parent_index = order_tree.selection()[0]
            val = []
            print(site_tree_removed)
            for a in order_tree.get_children():
                for b in order_tree.get_children(a):
                    val.append(order_tree.item(b)['values'])
            for count, a in enumerate(site_tree_removed):
                if count > len(val):
                    break
                site_tree_removed.remove(a)
                if check_var.get() == 0:
                    val[count].reverse()
                site_tree.insert('', 'end', a, values=val[count])
            order_tree.delete(parent_index)
            site_tree_index += len(val)
            removed_sections.append(parent_index)
            current_section = int(parent_index)

        def remove_section_site():
            nonlocal site_tree_index
            if len(order_tree.selection()) == 0:
                return
            index = order_tree.selection()[0]
            if ' ' not in index:
                return
            val = order_tree.item(index)['values']
            site_tree.insert('', 'end', site_tree_removed[0], values=val)
            site_tree_removed.pop(0)
            order_tree.delete(index)

        def sort_site():
            for a in order_tree.get_children():
                for b in order_tree.get_children(a):
                    val = order_tree.item(b)['values']
                    val.reverse()
                    order_tree.item(b, values=val)

        def context_popup(event):
            if len(order_tree.selection()) != 0:
                context.tk_popup(event.x_root, event.y_root, 0)

        def disable_enable(stage):
            if stage == 'start':
                option_1.configure(state='disabled')
                option_2.configure(state='disabled')
                option_3.configure(state='disabled')
                option_4.configure(state='disabled')
                op_var.set(0)
                op_var_2.set(0)
                load_btn.configure(state='disabled')
                rows_per_group_entry.configure(state='disabled')
                rows_per_group_label.configure(state='disabled')
                create_new_section_btn.configure(state='disabled')
                site_job_check.configure(state='disabled')
                site_tree.unbind('<Double-Button-1>')

            elif stage == 'mode':
                if var.get() == 'By Site':
                    option_1.configure(state='normal')
                    option_2.configure(state='normal')
                    option_3.configure(state='disabled')
                    option_4.configure(state='disabled')
                    op_var_2.set(0)
                    load_btn.configure(state='normal')
                    rows_per_group_entry.configure(state='disabled')
                    rows_per_group_label.configure(state='disabled')
                    create_new_section_btn.configure(state='disabled')
                    site_job_check.configure(state='disabled')
                    site_tree.unbind('<Double-Button-1>')
                elif var.get() == 'By Material':
                    option_1.configure(state='disabled')
                    option_2.configure(state='disabled')
                    load_btn.configure(state='disabled')
                    option_3.configure(state='normal')
                    option_4.configure(state='normal')
                    op_var.set(0)

            elif stage == 'set':
                rows_per_group_entry.configure(state='disabled')
                rows_per_group_label.configure(state='disabled')
                create_new_section_btn.configure(state='normal')
                site_job_check.configure(state='normal')
                site_tree.bind('<Double-Button-1>', lambda args: add_to_section())
                option_1.configure(state='normal')
                option_2.configure(state='normal')

            elif stage == 'auto':
                rows_per_group_entry.configure(state='normal')
                rows_per_group_label.configure(state='normal')
                load_btn.configure(state='normal')
                create_new_section_btn.configure(state='disabled')
                site_job_check.configure(state='disabled')
                site_tree.unbind('<Double-Button-1>')
                option_1.configure(state='normal')
                option_2.configure(state='normal')

            elif stage == 'ready':
                load_btn.configure(state='normal')

        def add_through_paste():
            paste_values_raw = self.app.root.clipboard_get()
            paste_values = paste_values_raw.split('\n')
            nonlocal site_tree_index
            parent_index = order_tree.selection()[0]
            index = int(order_row_index[int(order_tree.selection()[0]) - 1])
            for val in paste_values:
                val_2 = []
                for row in site_tree.get_children():
                    if site_tree.item(row)['values'][0] == val:
                        val_2 = site_tree.item(row)['values']
                        if check_var.get() == 0:
                            val_2.reverse()
                        site_tree_removed.append(row)
                        site_tree.delete(row)
                        site_tree_index -= 1

                if len(val_2) > 0:
                    order_tree.insert(parent_index, 'end', parent_index + ' ' + str(index), values=val_2)
                    index += 1

            index += 1
            order_row_index[int(order_tree.selection()[0]) - 1] = str(index)

        win = tk.Toplevel(self.app.root)
        wb = Workbook()
        ws = wb.active
        current_section = 0
        order_row_index = []
        site_tree_removed = []
        removed_sections = []

        var = tk.StringVar()
        var.set('Select Mode')
        option_frame = ttk.Frame(win, height=40)
        option_frame.pack_propagate(False)
        option_menu = ttk.OptionMenu(option_frame, var,
                                     var.get(),
                                     *['By Site', 'By Material'],
                                     command=lambda args: disable_enable('mode'))
        check_var = tk.IntVar()
        check_var.set(0)
        site_job_check = ttk.Checkbutton(win, text='Site ID', style='Switch.TCheckbutton',
                                         variable=check_var,
                                         onvalue=0,
                                         offvalue=1,
                                         command=sort_site)

        op_var = tk.IntVar()
        op_var.set(0)
        op_var_2 = tk.IntVar()
        op_var_2.set(0)
        option_1 = ttk.Radiobutton(win, text='1 per page', variable=op_var, value=1,
                                   command=lambda: disable_enable('ready'))
        option_2 = ttk.Radiobutton(win, text='X per page', variable=op_var, value=2,
                                   command=lambda: disable_enable('ready'))
        option_3 = ttk.Radiobutton(win, text='Set Print Order', variable=op_var_2, value=1,
                                   command=lambda: disable_enable('set'))
        frame_1 = ttk.Frame(win)
        option_4 = ttk.Radiobutton(frame_1, text='Automatic', variable=op_var_2, value=2,
                                   command=lambda: disable_enable('auto'))
        rows_per_group_entry = ttk.Entry(frame_1, width=4)
        rows_per_group_label = ttk.Label(frame_1, text='Rows per group')

        create_new_section_btn = ttk.Button(win, text='Create New Section', command=create_new_tree_section)

        bottom_frame = ttk.Frame(win)
        load_btn = ttk.Button(bottom_frame, text='Load', command=load_sheet)

        site_tree = ttk.Treeview(win)
        site_tree.bind('<Double-Button-1>', lambda args: add_to_section())
        site_tree.column('#0', width=0, stretch=False)
        site_tree['columns'] = [0, 1]
        site_tree.heading(0, text='Job#')
        site_tree.heading(1, text='Site ID')
        site_tree.column(0, width=120)
        site_tree.column(1, width=120)
        site_tree_index = 0
        for value in self.selected_sites:
            val_1 = [value, self.site_material[value][0][0]]
            site_tree.insert('', 'end', str(site_tree_index), values=val_1)
            site_tree_index += 1

        order_tree = ttk.Treeview(win)
        order_tree.bind('<Button-3>', context_popup)
        order_tree.bind('<Double-Button-1>', lambda args: remove_section_site())
        order_tree.column('#0', width=20, stretch=False)
        order_tree['columns'] = [0]
        order_tree.heading(0, text='Section')
        order_tree.column(0, width=200)

        context = tk.Menu(order_tree)
        context.add_command(label='Remove', command=remove_section)
        context.add_command(label='Paste', command=add_through_paste)

        create_new_section_btn.grid(row=3, column=0, sticky=tk.W, pady=(0, 5), padx=5)
        site_job_check.grid(row=3, column=1, sticky=tk.E, pady=(0, 5), padx=5)
        option_3.grid(row=1, column=0, sticky=tk.W, pady=(0, 5), padx=5)
        frame_1.grid(row=2, column=0, sticky=tk.W, pady=(0, 5), padx=5)

        option_4.grid(row=0, column=0, sticky=tk.W, pady=(0, 5), padx=5)
        rows_per_group_entry.grid(row=0, column=1, padx=(10, 5), pady=5)
        rows_per_group_entry.insert(0, '6')
        rows_per_group_label.grid(row=0, column=2, padx=5, pady=5, sticky=tk.W)

        site_tree.grid(row=4, column=0, padx=(5, 0), sticky=tk.W + tk.E)
        order_tree.grid(row=4, column=1, padx=(0, 5))

        option_frame.grid(row=0, column=0, columnspan=2, padx=5, pady=5, sticky=tk.W + tk.E)
        option_menu.pack(fill=tk.BOTH)
        option_1.grid(row=5, column=0, padx=5, pady=5, sticky=tk.W)
        option_2.grid(row=6, column=0, padx=5, pady=5, sticky=tk.W)

        bottom_frame.grid(row=7, column=0, columnspan=2, sticky=tk.E)
        load_btn.grid(row=0, column=1, padx=5, pady=5, sticky=tk.E)

        disable_enable('start')

    def translate(self, data):
        for key, value in self.translations_direct.items():
            if data[1] == key and data[0] == key:
                raise 'Huh?'
            elif data[0] == key:
                data[0] = value
            elif data[1] == key:
                data[1] = value

        for key, value in self.translations_indirect.items():
            if key in data[0] and key in data[1]:
                raise 'Huh?'
            elif key in data[0]:
                first_index_1 = data[0].find(key)
                first_index_2 = first_index_1 - int(value[0])
                first_part = data[0][first_index_1:first_index_2]
                second_index_1 = data[0].find(key)
                second_index_1 += len(key)
                second_index_2 = second_index_1 + int(value[2])
                second_part = data[0][second_index_1:second_index_2]
                data[0] = first_part + value[1] + second_part
            elif key in data[1]:
                first_index_1 = data[0].find(key)
                first_index_2 = first_index_1 - int(value[0])
                first_part = data[0][first_index_1:first_index_2]
                second_index_1 = data[0].find(key)
                second_index_1 += len(key)
                second_index_2 = second_index_1 + int(value[2])
                second_part = data[0][second_index_1:second_index_2]
                data[0] = first_part + value[1] + second_part

        return data

    def reload(self):
        clear_tree(self.import_tree_1, self.import_index_1)
        clear_tree(self.import_tree_2, self.import_index_2)
        self.import_index_1 = 0
        self.import_index_2 = 0
        self.import_tree(reload=True)


class DtobPage1:
    def __init__(self, app):
        self.app = app

        self.main_frame = ttk.Frame(self.app.root)
        self.frame_left = ttk.Frame(self.main_frame)
        self.frame_center = ttk.Frame(self.main_frame)
        self.frame_bottom = ttk.Frame(self.main_frame)
        self.frame_center_bot = ttk.Frame(self.frame_center)
        self.frame_center_top = ttk.Frame(self.frame_center)
        self.frame_right = ttk.Frame(self.main_frame)
        self.import_frame = ttk.Frame(self.frame_left)
        self.export_frame = ttk.Frame(self.frame_left)
        self.check_frame = ttk.Frame(self.frame_right)

        self.main_frame.pack(fill=tk.BOTH)
        self.frame_left.grid(row=0, column=0, sticky=tk.N + tk.S, rowspan=3)
        self.frame_center.grid(row=0, column=1)
        self.frame_center_top.grid(row=0, column=0, sticky=tk.N)
        self.frame_center_bot.grid(row=1, column=0, sticky=tk.S)
        self.frame_right.grid(row=0, column=2, sticky=tk.E + tk.N)
        self.frame_bottom.grid(row=2, column=2, sticky=tk.S + tk.E)
        self.import_frame.grid(row=0, column=0, sticky=tk.N)
        self.export_frame.grid(row=1, column=0, sticky=tk.S, pady=(0, 5))
        self.check_frame.grid(row=0, column=0, sticky=tk.N + tk.W)

        self.pad_y_left = (10, 0)
        self.frame_left.rowconfigure(1, weight=1)
        self.import_index = 0
        self.export_index = 0
        self.row_list = []
        self.filter_clicked = False

        self.export_dict = {}
        self.import_dict = {}
        self.export_delete = {}
        self.import_delete = {}
        self.filter_selections = []

        self.selected_col = None
        self.selected_import = False
        self.selected_export = False

        self.import_entry = ttk.Entry(self.import_frame)
        self.import_label = ttk.Label(self.import_frame, text='Import sheet')
        self.import_button = ttk.Button(self.import_frame, text='Enter', command=self.import_sheet)
        self.all_to_import_button = ttk.Button(self.import_frame, text='To Import', command=self.all_to_import)
        self.all_to_export_button = ttk.Button(self.import_frame, text='To Export', command=self.all_to_export)

        self.req_id_label = ttk.Label(self.export_frame, text='Request ID')
        self.req_id_entry = ttk.Entry(self.export_frame)
        self.material_id_label = ttk.Label(self.export_frame, text='Material ID')
        self.material_id_entry = ttk.Entry(self.export_frame)
        self.material_qty_label = ttk.Label(self.export_frame, text='Material Quantity')
        self.material_qty_entry = ttk.Entry(self.export_frame)
        self.to_market_label = ttk.Label(self.export_frame, text='Market')
        self.to_market_entry = ttk.Entry(self.export_frame)
        self.add_button = ttk.Button(self.export_frame, text='Add', command=self.add_items)

        self.import_tree = ttk.Treeview(self.frame_center_top)
        self.import_tree.bind('<Double-Button-1>', lambda args: self.import_to_export())
        self.import_tree['columns'] = ('0', '1', '2', '3')
        self.import_tree.column('#0', width=0, stretch=False)
        self.import_tree.heading('0', text='Request ID',
                                 command=lambda: self.select(heading='Request ID'))
        self.import_tree.heading('1', text='Material ID',
                                 command=lambda: self.select(heading='Material ID'))
        self.import_tree.heading('2', text='Material Quantity',
                                 command=lambda: self.select(heading='Material Quantity'))
        self.import_tree.heading('3', text='Market',
                                 command=lambda: self.select(heading='Market'))

        self.export_tree = ttk.Treeview(self.frame_center_bot)
        self.export_tree.bind('<Double-Button-1>', lambda args: self.export_to_import())
        self.export_tree['columns'] = ('0', '1', '2', '3')
        self.export_tree.column('#0', width=0, stretch=False)
        self.export_tree.heading('0', text='Request ID',
                                 command=lambda: self.select(heading='Request ID'))
        self.export_tree.heading('1', text='Material ID',
                                 command=lambda: self.select(heading='Material ID'))
        self.export_tree.heading('2', text='Material Quantity',
                                 command=lambda: self.select(heading='Material Quantity'))
        self.export_tree.heading('3', text='Market',
                                 command=lambda: self.select(heading='Market'))

        self.okay_button = ttk.Button(self.frame_bottom, text='Next', command=self.print_data)

        self.selected_import_check = ttk.Checkbutton(self.check_frame,
                                                     command=lambda: self.tree_select(self.selected_import_check))
        self.selected_export_check = ttk.Checkbutton(self.check_frame,
                                                     command=lambda: self.tree_select(self.selected_export_check))
        self.selected_import_label = ttk.Label(self.check_frame, text='Import Tree')
        self.selected_export_label = ttk.Label(self.check_frame, text='Export Tree')

        self.selected_label = ttk.Label(self.frame_right)
        self.filter_box = ttk.Combobox(self.frame_right)
        self.filter_box.bind('<Button-1>', lambda args: self.show_filter_options())
        self.filter_box.bind('<Return>', self.add_filter_value)
        self.filter_label = ttk.Label(self.frame_right, text='Filter')
        self.filter_enter = ttk.Button(self.frame_right, text='Enter', command=self.add_filter_value)
        self.filter_list = ttk.Treeview(self.frame_right, show='tree')
        self.filter_list['columns'] = ['value']
        self.filter_list.bind('<Double-Button-1>', lambda args: self.remove_filter_value())
        self.filter_list.column('#0', width=0, stretch=False)
        self.filter_list.column('value', width=0, stretch=True)
        self.filter_switch = ttk.Checkbutton(self.frame_right, style='Switch.TCheckbutton', command=self.filter_on_off)

        self.import_label.grid(row=0, column=0, sticky=tk.W + tk.E, pady=(20, 0))
        self.import_entry.grid(row=1, column=0, sticky=tk.W + tk.E, pady=self.pad_y_left)
        self.import_button.grid(row=2, column=0, sticky=tk.W + tk.E, pady=self.pad_y_left)
        self.all_to_import_button.grid(row=3, column=0, sticky=tk.W + tk.E, pady=(30, 0))
        self.all_to_export_button.grid(row=4, column=0, sticky=tk.W + tk.E, pady=(10, 0))

        self.req_id_label.grid(row=1, column=0, padx=5, pady=self.pad_y_left, sticky=tk.W + tk.E)
        self.req_id_entry.grid(row=2, column=0, padx=5, sticky=tk.W + tk.E)
        self.material_id_label.grid(row=3, column=0, padx=5, pady=self.pad_y_left, sticky=tk.W + tk.E)
        self.material_id_entry.grid(row=4, column=0, padx=5, sticky=tk.W + tk.E)
        self.material_qty_label.grid(row=5, column=0, padx=5, pady=self.pad_y_left, sticky=tk.W + tk.E)
        self.material_qty_entry.grid(row=6, column=0, padx=5, sticky=tk.W + tk.E)
        self.to_market_label.grid(row=7, column=0, padx=5, pady=self.pad_y_left, sticky=tk.W + tk.E)
        self.to_market_entry.grid(row=8, column=0, padx=5, sticky=tk.W + tk.E)
        self.add_button.grid(row=9, column=0, padx=5, pady=(20, 45), sticky=tk.W + tk.E)

        self.export_tree.pack(side='left', padx=(0, 10), pady=(5, 0))
        self.import_tree.pack(side='left', padx=(0, 10), pady=(10, 5))

        self.selected_import_check.grid(row=0, column=0, pady=(10, 5), padx=(0, 10), sticky=tk.W)
        self.selected_export_check.grid(row=1, column=0, pady=(5, 0), padx=(0, 10), sticky=tk.W)
        self.selected_import_label.grid(row=0, column=1, pady=(10, 5), padx=(5, 10), sticky=tk.W)
        self.selected_export_label.grid(row=1, column=1, pady=(5, 0), padx=(5, 10), sticky=tk.W)

        self.selected_label.grid(row=1, column=0, sticky=tk.E + tk.W, pady=(10, 0), padx=(0, 10))
        self.filter_label.grid(row=2, column=0, sticky=tk.E + tk.W, pady=(10, 0), padx=(0, 10))
        self.filter_box.grid(row=3, column=0, sticky=tk.E + tk.W, pady=(0, 5), padx=(0, 10))
        self.filter_enter.grid(row=4, column=0, sticky=tk.E + tk.W, pady=(5, 5), padx=(0, 10))
        self.filter_switch.grid(row=5, column=0, sticky=tk.E + tk.W, pady=(5, 5), padx=(0, 10))
        self.filter_list.grid(row=6, column=0, pady=(20, 0), sticky=tk.E + tk.W, padx=(0, 10))

        self.okay_button.grid(row=0, column=0, sticky=tk.E, padx=(0, 10), pady=(5, 5))

    def add_items(self):
        a_list = [self.req_id_entry.get(), self.material_id_entry.get(),
                  self.material_qty_entry.get(), self.to_market_entry.get()]

        self.export_tree.insert('', 'end', iid=str(self.export_index), values=a_list)
        self.export_dict.update({str(self.export_index): a_list})
        self.export_index += 1

    def select(self, heading=None):
        if heading == 'Request ID':
            self.selected_col = 0
        elif heading == 'Material ID':
            self.selected_col = 1
        elif heading == 'Material Quantity':
            self.selected_col = 2
        elif heading == 'Market':
            self.selected_col = 3

        self.selected_label.configure(text=heading)

    def print_data(self):
        row_count = 0
        all_borders = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        def initialize_wb():
            top_colors = PatternFill(patternType='solid', start_color='0070C0', end_color='0070C0')
            top_font = Font(color='FFFFFF')

            ws.cell(1, 1).value = 'Request ID'
            ws.cell(1, 2).value = 'Material'
            ws.cell(1, 3).value = 'Material QTY'
            ws.cell(1, 4).value = 'Market'
            for i in range(1, 5):
                ws.cell(1, i).border = all_borders
                ws.cell(1, i).fill = top_colors
                ws.cell(1, i).font = top_font

            ws.column_dimensions['A'].width = 13
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 13
            ws.column_dimensions['D'].width = 13

        def export_data():
            data_a = get_treeview_values(self.export_tree, col=0)
            data_b = get_treeview_values(self.export_tree, col=1)
            data_c = get_treeview_values(self.export_tree, col=2)
            data_d = get_treeview_values(self.export_tree, col=3)
            data = [data_a, data_b, data_c, data_d]
            nonlocal row_count
            row_count = len(data_a) + 1
            for a in range(len(data)):
                for c, b in enumerate(data[a]):
                    ws.cell(c + 2, a + 1).value = b

        def format_rows():
            for x in range(row_count + 1):
                ws.row_dimensions[x].height = 20

            for a in range(1, row_count):
                for b in range(4):
                    ws.cell(a + 1, b + 1).border = all_borders

        wb = Workbook()
        ws = wb.active
        initialize_wb()
        export_data()
        format_rows()
        ws.set_printer_settings(paper_size=90, orientation='landscape')
        wb.save(filename=r'transfer.xlsx')
        os.startfile(r'transfer.xlsx')
        self.next_page()

    def next_page(self):
        def prepare_data():
            markets = get_treeview_values(self.export_tree, col=3)
            markets = dict.fromkeys(markets)

            for market, value in markets.items():
                for index in self.export_tree.get_children():
                    tree_market = self.export_tree.item(index)['values'][3]
                    if tree_market == market:
                        tree_val = [
                            self.export_tree.item(index)['values'][0],
                            self.export_tree.item(index)['values'][1],
                            self.export_tree.item(index)['values'][2],
                        ]
                        if isinstance(value, list):
                            value.append(tree_val)
                        elif value is None:
                            value = tree_val

            return markets

        self.main_frame.pack_forget()
        data = prepare_data()
        self.app.page = DtobPage2(self.app, data)

    def import_sheet(self):
        path = self.import_entry.get()
        wb = load_workbook(filename=path)
        ws = wb.active

        last_row = ws.max_row

        for x in range(2, last_row + 1):
            # 1 = req_id, 14 = market, 9 = material, 11 = qty
            if ws.row_dimensions[x].hidden is False:
                row = [ws.cell(x, 1).value, ws.cell(x, 9).value, ws.cell(x, 11).value, ws.cell(x, 14).value]
                for i in range(len(row)):
                    if isinstance(row[i], int):
                        row[i] = str(i)
                    elif isinstance(row[i], float):
                        row[i] = int(row[i])
                        row[i] = str(row[i])
                self.row_list.append(row)

        for rows in self.row_list:
            self.import_tree.insert('', 'end', str(self.import_index), values=rows)
            self.import_delete.update({str(self.import_index): rows})
            self.export_delete.update({str(self.import_index): rows})
            self.import_index += 1
            self.export_index += 1

    def filter_on(self):
        if len(self.filter_selections) > 0:
            if self.selected_import is True:
                for index in self.import_tree.get_children():
                    values = self.import_tree.item(index)['values']
                    for filter_column, filter_value in self.filter_selections:
                        if str(values[filter_column]) == str(filter_value):
                            break
                    else:
                        self.import_dict.update({str(index): values})
                        self.import_tree.detach(str(index))

            if self.selected_export is True:
                for index in self.export_tree.get_children():
                    values = self.export_tree.item(index)['values']
                    for filter_column, filter_value in self.filter_selections:
                        if str(values[filter_column]) == str(filter_value):
                            break

                    else:
                        self.export_dict.update({str(index): values})
                        self.export_tree.detach(str(index))

    def add_filter_value(self):
        self.filter_list.insert('', 'end', str(self.filter_box.get()), values=[str(self.filter_box.get()), ])
        value = str(self.filter_box.get())
        column = self.selected_col
        selected = [column, value]
        self.filter_selections.append(selected)
        self.filter_box.delete(0, 'end')
        if self.filter_clicked is True:
            self.filter_on()

    def show_filter_options(self):
        if self.selected_col is not None:
            self.filter_box['values'] = ()
            for row in self.row_list:
                if self.filter_box.get() in row[self.selected_col]:
                    if row[self.selected_col] not in str(self.filter_selections):
                        if row[self.selected_col] not in self.filter_box['values']:
                            if isinstance(self.filter_box['values'], str):
                                self.filter_box['values'] = (self.filter_box['values'], row[self.selected_col])
                            else:
                                self.filter_box['values'] += (row[self.selected_col],)

    def remove_filter_value(self):
        if not len(self.filter_list.selection()) == 0:
            value = self.filter_list.selection()[0]
            index = self.filter_list.index(value)
            self.filter_list.delete(value)
            self.filter_selections.pop(int(index))
            if self.filter_clicked is True:
                self.filter_on()

    def filter_on_off(self):
        if self.filter_clicked is False:
            self.filter_clicked = True
            self.filter_on()
        elif self.filter_clicked is True:
            self.filter_clicked = False
            self.filter_off()

    def filter_off(self):
        for index, value in self.import_dict.items():
            for tree_index in self.export_tree.get_children():
                values = self.export_tree.item(tree_index)['values']
                for x in range(len(values)):
                    values[x] = str(values[x])

                if values == value:
                    break
            else:
                self.import_tree.reattach(parent='', item=str(index), index=int(index))

        for index, value in self.export_dict.items():
            for tree_index in self.import_tree.get_children():
                values = self.import_tree.item(tree_index)['values']
                for x in range(len(values)):
                    values[x] = str(values[x])

                if values == value:
                    break
            else:
                self.export_tree.reattach(parent='', item=str(index), index=int(index))

    def tree_select(self, selected_check):
        if selected_check is self.selected_import_check:
            if self.selected_import is False:
                self.selected_import = True
            elif self.selected_import is True:
                self.selected_import = False
        elif selected_check is self.selected_export_check:
            if self.selected_export is False:
                self.selected_export = True
            elif self.selected_export is True:
                self.selected_export = False

        if self.filter_clicked is True:
            self.filter_on()

    def export_to_import(self):
        if len(self.export_tree.selection()) != 0:
            index = self.export_tree.selection()[0]
            values = self.export_tree.item(index)['values']

            for key, value in self.export_dict.items():
                if value == values:
                    export_dict_index = key
                    self.export_dict.pop(export_dict_index)
                    break

            for x in range(len(values)):
                values[x] = str(values[x])

            for key, value in self.import_delete.items():
                if value == values:
                    import_delete_index = key
                    break
            else:
                raise ValueError('import_delete_index not found')

            self.export_tree.delete(index)
            self.import_tree.insert('', int(import_delete_index), str(import_delete_index), values=values)

    def import_to_export(self):
        if len(self.import_tree.selection()) != 0:
            index = self.import_tree.selection()[0]
            values = self.import_tree.item(index)['values']

            for key, value in self.import_dict.items():
                if value == values:
                    pop_index = key
                    self.import_dict.pop(pop_index)
                    break

            for x in range(len(values)):
                values[x] = str(values[x])

            for key, value in self.export_delete.items():
                if value == values:
                    export_delete_index = key
                    break
            else:
                raise ValueError('export_delete_index not found')

            self.import_tree.delete(index)
            self.export_tree.insert('', 'end', str(export_delete_index), values=values)

    def all_to_export(self):
        for index in self.import_tree.get_children():
            values = self.import_tree.item(index)['values']

            for key, value in self.import_dict.items():
                if value == values:
                    self.import_dict.pop(key)
                    break

            self.import_tree.delete(index)
            self.export_tree.insert('', 'end', str(index), values=values)

    def all_to_import(self):
        for index in self.export_tree.get_children():
            values = self.export_tree.item(index)['values']

            for key, value in self.export_dict.items():
                if value == values:
                    self.export_dict.pop(key)
                    break

            self.export_tree.delete(index)
            self.import_tree.insert('', 'end', str(index), values=values)


class DtobPage2:
    class ImportTree:
        def __init__(self, import_frame, market_data):
            self.tree = ttk.Treeview(import_frame)
            self.tree['columns'] = [0, 1, 2]
            self.tree.column('#0', width=0, stretch=False)
            self.tree.heading(0, text='Request ID')
            self.tree.heading(1, text='Material ID')
            self.tree.heading(2, text='Material Quantity')
            self.tree.pack()
            if market_data is None:
                pass

    class ExportTree:
        def __init__(self, export_frame, container):
            self.tree = ttk.Treeview(export_frame)
            self.tree['columns'] = [0, 1, 2]
            self.tree.column('#0', width=0, stretch=False)
            self.tree.heading(0, text='Request ID')
            self.tree.heading(1, text='Material ID')
            self.tree.heading(2, text='Material Quantity')
            self.tree.pack()

    class ContainerTree:
        def __init__(self, container_frame, market):
            self.tree = ttk.Treeview(container_frame, show='tree')
            self.tree['columns'] = [0]
            self.tree.column('#0', width=0, stretch=False)
            self.tree.column(0, width=80, stretch=True)
            self.tree.pack()

    def __init__(self, app, data):
        self.app = app
        self.market_index = 0
        self.import_frame = ttk.Frame(self.app.root)
        self.export_frame = ttk.Frame(self.app.root)
        self.market_frame = ttk.Frame(self.app.root)
        self.container_frame = ttk.Notebook(self.app.root)
        self.control_frame = ttk.Frame(self.app.root)

        self.market_frame.grid(row=0, column=0, sticky=tk.N + tk.S, padx=(10, 0), pady=(10, 0))
        self.import_frame.grid(row=0, column=1, padx=(0, 10), pady=(10, 0))
        self.control_frame.grid(row=1, column=0, columnspan=2, sticky=tk.W + tk.E, padx=10)
        self.container_frame.grid(row=2, column=0, sticky=tk.N + tk.S, padx=(10, 0), pady=(0, 10))
        self.export_frame.grid(row=2, column=1, padx=(0, 10), pady=(0, 10))

        self.market_list = ttk.Treeview(self.market_frame)
        self.market_list['columns'] = [0]
        self.market_list.column('#0', width=0, stretch=False)
        self.market_list.column(0, width=80, stretch=True)
        self.market_list.heading(0, text='Market')
        for key in data.keys():
            self.market_list.insert('', 'end', str(self.market_index), values=key)
            self.market_index += 1
        self.market_list.pack(fill='y', expand=True)

        self.add_button = ttk.Button(self.control_frame, text='Add', command=self.add_con_row)
        self.add_button.grid(row=0, column=0, padx=1, pady=1)

        self.type_1 = ttk.Frame(self.container_frame)
        self.type_2 = ttk.Frame(self.container_frame)
        self.type_1.pack()
        self.type_2.pack()
        self.container_frame.add(self.type_1, text='Pallets')
        self.container_frame.add(self.type_2, text='Boxes')

        self.container_list_1 = self.ContainerTree(self.type_1, None)
        self.container_list_2 = self.ContainerTree(self.type_2, None)
        self.import_tree = self.ImportTree(self.import_frame, None)
        self.export_tree = self.ExportTree(self.export_frame, None)

    def add_con_row(self):
        pass


startapp = App()
